#!/usr/bin/env python3
"""
Edit Excel files.

Usage:
    python edit_xlsx.py --file input.xlsx --output output.xlsx --operations '[...]'
"""

import argparse
import sys
import json
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
except ImportError:
    print("Error: 'openpyxl' package is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def apply_operation(wb, operation):
    """Apply a single edit operation."""
    action = operation.get('action')
    sheet_name = operation.get('sheet', wb.active.title)
    
    if action == 'set_cell':
        ws = wb[sheet_name]
        cell = operation.get('cell', 'A1')
        value = operation.get('value')
        ws[cell] = value
    
    elif action == 'set_range':
        ws = wb[sheet_name]
        start = operation.get('start', 'A1')
        data = operation.get('data', [])
        
        # Parse start cell
        col_letter = ''.join(filter(str.isalpha, start))
        row_num = int(''.join(filter(str.isdigit, start)))
        
        for row_idx, row_data in enumerate(data):
            for col_idx, value in enumerate(row_data):
                ws.cell(row=row_num + row_idx, column=ord(col_letter) - ord('A') + 1 + col_idx, value=value)
    
    elif action == 'add_formula':
        ws = wb[sheet_name]
        cell = operation.get('cell', 'A1')
        formula = operation.get('formula', '')
        ws[cell] = formula
    
    elif action == 'add_sheet':
        name = operation.get('name', 'NewSheet')
        wb.create_sheet(title=name)
    
    elif action == 'format_cell':
        ws = wb[sheet_name]
        cell = operation.get('cell', 'A1')
        cell_obj = ws[cell]
        
        if operation.get('bold'):
            cell_obj.font = Font(bold=True)
        if operation.get('bg_color'):
            cell_obj.fill = PatternFill(start_color=operation['bg_color'], 
                                         end_color=operation['bg_color'], 
                                         fill_type='solid')


def edit_xlsx(input_path: str, output_path: str, operations: list):
    """Edit an Excel file."""
    wb = load_workbook(input_path)
    
    for operation in operations:
        apply_operation(wb, operation)
    
    wb.save(output_path)
    return output_path


def main():
    parser = argparse.ArgumentParser(description='Edit Excel files')
    parser.add_argument('--file', required=True, help='Input Excel file')
    parser.add_argument('--output', required=True, help='Output file path')
    parser.add_argument('--operations', required=True, help='JSON array of operations')
    
    args = parser.parse_args()
    
    if not Path(args.file).exists():
        print(f"Error: File not found: {args.file}", file=sys.stderr)
        sys.exit(1)
    
    try:
        operations = json.loads(args.operations)
    except json.JSONDecodeError as e:
        print(f"Error: Invalid operations JSON - {e}", file=sys.stderr)
        sys.exit(1)
    
    try:
        result = edit_xlsx(args.file, args.output, operations)
        print(json.dumps({"status": "success", "file": result}))
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == '__main__':
    main()
