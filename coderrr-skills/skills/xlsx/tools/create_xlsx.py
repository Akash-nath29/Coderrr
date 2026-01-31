#!/usr/bin/env python3
"""
Create Excel workbooks.

Usage:
    python create_xlsx.py --output data.xlsx --sheets '[...]'
"""

import argparse
import sys
import json
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: 'openpyxl' package is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def create_xlsx(output_path: str, sheets_spec: list):
    """Create an Excel workbook."""
    wb = Workbook()
    
    # Remove default sheet if we're creating new ones
    if sheets_spec:
        default_sheet = wb.active
        wb.remove(default_sheet)
    
    for sheet_spec in sheets_spec:
        sheet_name = sheet_spec.get('name', 'Sheet1')
        ws = wb.create_sheet(title=sheet_name)
        
        row_num = 1
        
        # Add headers
        headers = sheet_spec.get('headers', [])
        if headers:
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col, value=header)
                cell.font = Font(bold=True)
            row_num += 1
        
        # Add data
        data = sheet_spec.get('data', [])
        for row_data in data:
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col)
                if isinstance(value, str) and value.startswith('='):
                    cell.value = value  # Formula
                else:
                    cell.value = value
            row_num += 1
        
        # Set column widths
        col_widths = sheet_spec.get('column_widths', {})
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width
    
    wb.save(output_path)
    return output_path


def main():
    parser = argparse.ArgumentParser(description='Create Excel workbooks')
    parser.add_argument('--output', required=True, help='Output file path (.xlsx)')
    parser.add_argument('--sheets', required=True, help='JSON specification of sheets')
    
    args = parser.parse_args()
    
    try:
        sheets = json.loads(args.sheets)
    except json.JSONDecodeError as e:
        print(f"Error: Invalid sheets JSON - {e}", file=sys.stderr)
        sys.exit(1)
    
    try:
        result = create_xlsx(args.output, sheets)
        print(json.dumps({"status": "success", "file": result, "sheets": len(sheets)}))
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == '__main__':
    main()
