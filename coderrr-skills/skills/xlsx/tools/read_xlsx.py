#!/usr/bin/env python3
"""
Read data from Excel files.

Usage:
    python read_xlsx.py --file data.xlsx --format json
"""

import argparse
import sys
import json
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: 'openpyxl' package is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def parse_range(range_str: str):
    """Parse range like A1:C10 into start and end."""
    if ':' in range_str:
        start, end = range_str.split(':')
        return start, end
    return range_str, range_str


def read_xlsx(file_path: str, sheet_name: str = None, cell_range: str = None, output_format: str = 'json'):
    """Read Excel file data."""
    wb = load_workbook(file_path, data_only=True)
    
    # Get sheet
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        ws = wb[sheet_name]
    else:
        ws = wb.active
    
    # Determine range
    if cell_range:
        data = []
        for row in ws[cell_range]:
            row_data = [cell.value for cell in row]
            data.append(row_data)
    else:
        data = []
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                data.append(list(row))
    
    # Format output
    if output_format == 'json':
        return json.dumps({
            "sheet": ws.title,
            "range": cell_range or f"A1:{get_column_letter(ws.max_column)}{ws.max_row}",
            "data": data
        }, indent=2, default=str)
    
    elif output_format == 'csv':
        lines = []
        for row in data:
            line = ','.join(str(cell) if cell is not None else '' for cell in row)
            lines.append(line)
        return '\n'.join(lines)
    
    else:  # text
        lines = []
        for row in data:
            line = '\t'.join(str(cell) if cell is not None else '' for cell in row)
            lines.append(line)
        return '\n'.join(lines)


def main():
    parser = argparse.ArgumentParser(description='Read Excel files')
    parser.add_argument('--file', required=True, help='Path to Excel file')
    parser.add_argument('--sheet', help='Sheet name')
    parser.add_argument('--range', help='Cell range (e.g., A1:D10)')
    parser.add_argument('--format', choices=['json', 'csv', 'text'], default='json')
    
    args = parser.parse_args()
    
    if not Path(args.file).exists():
        print(f"Error: File not found: {args.file}", file=sys.stderr)
        sys.exit(1)
    
    try:
        result = read_xlsx(args.file, args.sheet, args.range, args.format)
        print(result)
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == '__main__':
    main()
